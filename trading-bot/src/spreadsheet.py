import threading
from datetime import datetime
from pathlib import Path

SPREADSHEET_PATH = Path("/mnt/t7/trading-toolkit/Trading_Toolkit.xlsx")
BACKUP_PATH = Path("/home/nick/trading-toolkit/Trading_Toolkit.xlsx")
_write_lock = threading.Lock()


def _get_path() -> Path:
    if SPREADSHEET_PATH.exists():
        return SPREADSHEET_PATH
    if BACKUP_PATH.exists():
        return BACKUP_PATH
    raise FileNotFoundError("Trading_Toolkit.xlsx not found.")


def log_trade(symbol: str, action: str, entry_price: float, exit_price: float | None,
              qty: float, pnl: float | None = None, strategy: str = "RSI+MA", notes: str = ""):
    def _write():
        with _write_lock:
            try:
                from openpyxl import load_workbook
                path = _get_path()
                wb = load_workbook(str(path))
                ws = wb["Trade Log"]
                r = 5
                while ws.cell(r, 1).value is not None and r < 1000:
                    r += 1
                ws.cell(r, 1).value = datetime.now()
                ws.cell(r, 2).value = symbol.replace("/USD", "")
                ws.cell(r, 3).value = action
                if entry_price:
                    ws.cell(r, 4).value = round(entry_price, 4)
                if exit_price:
                    ws.cell(r, 5).value = round(exit_price, 4)
                ws.cell(r, 6).value = round(qty, 4) if qty else ""
                ws.cell(r, 13).value = strategy
                if notes:
                    ws.cell(r, 14).value = notes
                wb.save(str(path))
                print(f"  [SHEET] Logged {action} {symbol} row {r}")
            except Exception as e:
                print(f"  [SHEET] Failed to log trade: {e}")
    threading.Thread(target=_write, daemon=True).start()


def log_buy(symbol: str, entry_price: float, qty: float, strategy: str = "RSI+MA"):
    log_trade(symbol, "BUY", entry_price, None, qty, None, strategy)


def log_sell(symbol: str, action: str, entry_price: float, exit_price: float,
             qty: float, pnl: float, strategy: str = "RSI+MA"):
    log_trade(symbol, action, entry_price, exit_price, qty, pnl, strategy)
